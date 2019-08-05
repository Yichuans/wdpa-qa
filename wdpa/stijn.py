# import package
import os
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows    
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle

def output_errors_to_excel(result, outpath, checks, datatype):
    '''
    The functions_list is a list that contains all the names of the 
    functions (tests) of the WDPA QA. If the function's name is present
    in the result dictionary produced by the main function, 
    the WDPA failed the test and contains errors.
    
    Function output_errors_to_excel evaluates whether each function's name 
    in functions_list, is present in the result dictionary. 
    If so, it will output the coresponding DataFrame to a sheet in an 
    Excel file. In the Excel Summary sheet, the function's name will be 
    added along with string 'Fail'. Else, the test's name will be added to 
    the Excel Summary sheet along with string 'Pass'.
        
    ## Arguments ##
    result --         dictionary created by the main function, containing 
                      functions' names and DataFrames for tests that failed. 
                      In this dictionary, a function's name is the 'key', and 
                      the DataFrame with errors is the 'value'.
    
    outpath --        the output directory where the Excel file is to be saved
    
    checks --         a dictionary containing all the descriptive names 
                      and function names of the WDPA QA checks
    
    datatype --       a string specifying the input type: e.g. point or poly
                      This will be added to the Excel file's name.

    ## Example ##
    output_errors_to_excel(result=result,
                           outpath='C:\\Users\\paintern\\Desktop\\Stijn\\3. Data\\Test data',
                           checks=poly_checks,
                           datatype='poly'])
    '''
        
    # Function to find the row of the function_name
    # Required to add a hyperlink of the tab to the 
    # corresponding function_name cell
    def find_row(function_name):
        for row in wb['Summary'].iter_rows(min_row=1, max_row=wb['Summary'].max_row, max_col=1):
            for cell in row: # iterate through the cells in rows 1 through the end
                if cell.value == function_name: # select the correct cell
                    return (wb['Summary'].cell(row=cell.row, column=1).row) # return cell's row number
    
    # Set variables - to later add the current day to the filename
    filename = f'{datetime.datetime.now().strftime("%d%b%Y")}_WDPA_QA_checks_{datatype}.xlsx'
    output = outpath + os.sep + filename
    
    # Create the Excel workbook and the Summary sheet
    wb = Workbook()
    wb['Sheet'].title = 'Summary' # change default sheet's title
    wb["Summary"].append(["CHECK","RESULT", "COUNT"]) # add header for Summary sheet

    # If the function's name - in the functions_list - is present in the 
    # result dictionary, add DataFrame to a new sheet
    function_names = [each['name'] for each in checks] # make a list of all checks' names

    for function_name in function_names:
        if function_name in result:
            ws = wb.create_sheet(function_name)
        # export DataFrame rows to Excel
            for row in dataframe_to_rows(result[function_name], index=False): 
                ws.append(row)
        # Add a hyperlink to each sheet, to return to the Summary with a single click
            ws.insert_cols(1) # insert column at first position
            ws.cell(row=1, column=1).value = 'To Summary'
            ws.cell(row=1, column=1).hyperlink = (f'#Summary!A1')
            ws.cell(row=1, column=1).style = 'Hyperlink'
            ws.column_dimensions['A'].width = 14 # adjust width of column A
            ws.freeze_panes = 'B2'
        # add 'Check' or 'Fail' to Summary sheet
            if not function_name.startswith('ivd'):
                wb['Summary'].append([function_name,'Check', len(result[function_name])])
                ws.sheet_properties.tabColor = '87CEFA' # blue tab
                link = f'#{function_name}!A1' # create link to cell A1 of function_name tab
                # add link to the function_name of the Summary sheet, with hyperlink style
                wb['Summary'].cell(row=find_row(function_name), column=1).hyperlink = link 
                wb['Summary'].cell(row=find_row(function_name), column=1).style = 'Hyperlink'
            else:
                wb['Summary'].append([function_name,'Fail', len(result[function_name])])
                ws.sheet_properties.tabColor = 'F08080' # red tab
                link = f'#{function_name}!A1' # as above
                wb['Summary'].cell(row=find_row(function_name), column=1).hyperlink = link
                wb['Summary'].cell(row=find_row(function_name), column=1).style = 'Hyperlink'
        # add 'Pass' to Summary sheet as no rows with invalid WDPA_PIDs are present
        else:
            wb['Summary'].append([function_name,'Pass'])

    # Conditional formatting - different colours for Check, Fail, and Pass
    def add_conditional_formatting(colour, summary_result, sheetname):
        '''
        Add conditional formatting in the Summary sheet in Excel, for each check performed.
        This will fill cells containing certain values. 
        Red for failed tests, blue for tests that need to be checked, green for passed tests.

        ## Arguments ##
        colour --         A string specifying hex colour code (RRGGBB) to use for filling.
        summary_result -- A string with the summary result, either 'Fail', 'Check', or 'Pass'
        sheetname --      A string specifying the target sheet for the formatting, i.e. 'Summary'

        ## Example ##
        add_conditional_formatting(colour='87CEFA',
        summary_result='Check',
        sheetname='Summary)
        '''
        
        fill_col = PatternFill(bgColor=colour) # specify colour
        style_to_apply = DifferentialStyle(fill=fill_col) # specifyl style (fill)
        r = Rule(type="expression", dxf=style_to_apply, stopIfTrue=True) # specify rule
        r.formula = [f'$B2="{summary_result}"'] # only search in Column B, starting on second row
        wb[sheetname].conditional_formatting.add(f'A2:C{wb[sheetname].max_row}', r) # apply formatting

    add_conditional_formatting('87CEFA', 'Check', 'Summary') # blue
    add_conditional_formatting('F08080', 'Fail', 'Summary') # red
    add_conditional_formatting('98FB98', 'Pass', 'Summary') # green
    
    # Extra formatting
    wb['Summary'].sheet_properties.tabColor = '000000' # black tab 
    wb['Summary'].column_dimensions['A'].width = 31 # adjust column A's width
    wb['Summary'].freeze_panes = 'A2' # freeze header

    # Save the workbook
    wb.save(output)
    return